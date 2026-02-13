from __future__ import annotations

import io
import os
import re
import subprocess
import zipfile
from datetime import datetime
from typing import Any
from urllib.parse import urljoin, urlparse

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

from ercot_queue.config import (
    DEFAULT_DATA_PRODUCT_URLS,
    DEFAULT_REPORT_INDEX_URLS,
    REQUEST_TIMEOUT_SECONDS,
)

DATA_EXTENSIONS = (".csv", ".xls", ".xlsx", ".zip")
DATE_PATTERNS = [
    re.compile(r"(20\d{2})[-_/](\d{1,2})[-_/](\d{1,2})"),
    re.compile(r"(\d{1,2})[-_/](\d{1,2})[-_/](20\d{2})"),
    re.compile(r"(20\d{2})(\d{2})(\d{2})"),
]
MONTH_NAME_PATTERN = re.compile(
    r"(jan(?:uary)?|feb(?:ruary)?|mar(?:ch)?|apr(?:il)?|may|jun(?:e)?|jul(?:y)?|aug(?:ust)?|"
    r"sep(?:t(?:ember)?)?|oct(?:ober)?|nov(?:ember)?|dec(?:ember)?)[-_ ]*(20\d{2})",
    flags=re.IGNORECASE,
)
MONTH_TO_INT = {
    "jan": 1,
    "feb": 2,
    "mar": 3,
    "apr": 4,
    "may": 5,
    "jun": 6,
    "jul": 7,
    "aug": 8,
    "sep": 9,
    "oct": 10,
    "nov": 11,
    "dec": 12,
}
MAX_HEADER_SCAN_ROWS = 60
DEFAULT_OPERATING_ASSETS_INDEX_URL = "https://www.ercot.com/gridinfo/resource"


def fetch_latest_ercot_queue(custom_url: str | None = None) -> tuple[pd.DataFrame, dict[str, Any]]:
    """Fetches ERCOT interconnection queue data from ERCOT-hosted sources only."""
    errors: list[str] = []

    if custom_url:
        _assert_ercot_url(custom_url, "Custom URL")
        if "data-product-details" in custom_url:
            index_url, product_meta = discover_report_index_from_product_page(custom_url)
            report_url = discover_latest_report_url(index_url)
            df = _download_table(report_url)
            base_meta = {
                "source": "ercot_data_product",
                "source_url": report_url,
                "index_url": index_url,
                **product_meta,
            }
            source_meta = _attach_tab_metadata(df, base_meta)
        else:
            df = _download_table(custom_url)
            source_meta = _attach_tab_metadata(df, {"source": "ercot_custom", "source_url": custom_url})

        if df.empty:
            raise RuntimeError(f"Custom ERCOT URL returned no rows: {custom_url}")
        return df, source_meta

    product_urls = list(DEFAULT_DATA_PRODUCT_URLS)
    env_product_url = os.environ.get("ERCOT_DATA_PRODUCT_URL")
    if env_product_url:
        _assert_ercot_url(env_product_url, "ERCOT_DATA_PRODUCT_URL")
        product_urls.insert(0, env_product_url)

    for product_url in product_urls:
        try:
            _assert_ercot_url(product_url, "ERCOT data product URL")
            index_url, product_meta = discover_report_index_from_product_page(product_url)
            _assert_ercot_url(index_url, "Derived ERCOT index URL")
            report_url = discover_latest_report_url(index_url)
            _assert_ercot_url(report_url, "Discovered report URL")

            df = _download_table(report_url)
            if not df.empty:
                return df, _attach_tab_metadata(df, {
                    "source": "ercot_data_product",
                    "source_url": report_url,
                    "index_url": index_url,
                    **product_meta,
                })
        except Exception as exc:  # pylint: disable=broad-except
            errors.append(f"{product_url}: {exc}")

    index_urls = list(DEFAULT_REPORT_INDEX_URLS)
    env_index_url = os.environ.get("ERCOT_REPORT_INDEX_URL")
    if env_index_url:
        _assert_ercot_url(env_index_url, "ERCOT_REPORT_INDEX_URL")
        index_urls.insert(0, env_index_url)

    for index_url in index_urls:
        try:
            _assert_ercot_url(index_url, "ERCOT index URL")
            report_url = discover_latest_report_url(index_url)
            _assert_ercot_url(report_url, "Discovered report URL")

            df = _download_table(report_url)
            if not df.empty:
                return df, _attach_tab_metadata(df, {
                    "source": "ercot_mis",
                    "source_url": report_url,
                    "index_url": index_url,
                })
        except Exception as exc:  # pylint: disable=broad-except
            errors.append(f"{index_url}: {exc}")

    error_body = " | ".join(errors) if errors else "No ERCOT source returned data."
    raise RuntimeError(
        "Could not fetch ERCOT interconnection queue data from ERCOT-hosted sources. "
        "Set a direct ERCOT file URL in the app sidebar and try again. Details: "
        f"{error_body}"
    )


def fetch_latest_operating_assets(index_url: str | None = None) -> tuple[pd.DataFrame, dict[str, Any]]:
    """Fetch latest ERCOT operating assets sheet from the MORA workbook."""
    resource_index_url = (index_url or DEFAULT_OPERATING_ASSETS_INDEX_URL).strip()
    _assert_ercot_url(resource_index_url, "ERCOT operating assets index URL")

    workbook_url, workbook_label = discover_latest_mora_workbook_url(resource_index_url)
    content, _, effective_url = _fetch_url(workbook_url)
    xls = pd.ExcelFile(io.BytesIO(content))

    preferred_sheet = _select_operating_assets_sheet(xls.sheet_names)
    sheet_df, header_row, header_score = _read_sheet_best_effort(xls, preferred_sheet)
    cleaned = sheet_df.dropna(axis=1, how="all").dropna(axis=0, how="all")
    if cleaned.empty:
        raise RuntimeError(f"Latest MORA workbook sheet is empty: {preferred_sheet}")

    cleaned["_source_sheet"] = preferred_sheet
    cleaned["_source_header_row"] = header_row
    cleaned["_source_header_score"] = header_score

    meta = {
        "source": "ercot_mora",
        "source_url": effective_url,
        "index_url": resource_index_url,
        "report_label": workbook_label,
        "tab_count": len(xls.sheet_names),
        "tabs_processed": [preferred_sheet],
    }
    return cleaned, meta


def discover_latest_mora_workbook_url(index_url: str) -> tuple[str, str]:
    content, _, effective_url = _fetch_url(index_url)
    html_text = content.decode("utf-8", errors="ignore")
    soup = BeautifulSoup(html_text, "html.parser")

    candidates: list[dict[str, Any]] = []
    position = 0
    for anchor in soup.find_all("a"):
        href = (anchor.get("href") or "").strip()
        if not href:
            continue
        label = " ".join(anchor.get_text(" ", strip=True).split())
        abs_url = urljoin(effective_url, href)
        lower_url = abs_url.lower()
        lower_label = label.lower()

        if not _is_ercot_url(abs_url):
            continue
        if not lower_url.endswith(".xlsx"):
            continue
        if "mora" not in lower_url and "mora" not in lower_label:
            continue

        found_date = _parse_date(abs_url) or _parse_date(label) or datetime.min
        score = 0
        if "mora" in lower_url:
            score += 10
        if "mora" in lower_label:
            score += 8
        if "resource" in lower_label:
            score += 2

        candidates.append(
            {
                "url": abs_url,
                "label": label,
                "date": found_date,
                "score": score,
                "position": position,
            }
        )
        position += 1

    if not candidates:
        raise RuntimeError(f"No ERCOT MORA XLSX links found on page: {index_url}")

    candidates.sort(key=lambda item: (item["date"], item["score"], -item["position"]), reverse=True)
    best = candidates[0]
    return str(best["url"]), str(best["label"] or "MORA workbook")


def _select_operating_assets_sheet(sheet_names: list[str]) -> str:
    if not sheet_names:
        raise RuntimeError("MORA workbook has no sheets")

    preferred_patterns = [
        r"resource\s*details?",
        r"resource\s*detail",
        r"operating",
    ]
    lowered = [(name, name.lower()) for name in sheet_names]
    for pattern in preferred_patterns:
        for original, lower in lowered:
            if re.search(pattern, lower):
                return original

    return sheet_names[0]


def discover_report_index_from_product_page(product_url: str) -> tuple[str, dict[str, Any]]:
    content, _, effective_url = _fetch_url(product_url)
    html_text = content.decode("utf-8", errors="ignore")
    soup = BeautifulSoup(html_text, "html.parser")
    text = soup.get_text(" ", strip=True)

    report_type_match = re.search(r"report\s*type\s*id\s*(\d+)", text, flags=re.IGNORECASE)
    if not report_type_match:
        report_type_match = re.search(r"reportTypeId=(\d+)", html_text, flags=re.IGNORECASE)
    if not report_type_match:
        raise RuntimeError("Could not find Report Type ID on ERCOT data product page")

    report_type_id = report_type_match.group(1)
    index_url = (
        "https://www.ercot.com/misapp/GetReports.do"
        f"?reportTypeId={report_type_id}&showHTMLView=&mimicKey"
    )

    return index_url, {
        "data_product_url": effective_url,
        "report_type_id": report_type_id,
    }


def discover_latest_report_url(index_url: str) -> str:
    content, _, effective_url = _fetch_url(index_url)
    html_text = content.decode("utf-8", errors="ignore")

    candidates = _extract_report_candidates(html_text, effective_url)
    candidates = [candidate for candidate in candidates if _is_ercot_url(str(candidate["url"]))]
    if not candidates:
        raise RuntimeError("No ERCOT report links found on ERCOT index page")

    return str(_select_best_candidate(candidates)["url"])


def fetch_summary_under_study_mw(report_url: str) -> float | None:
    """Fetch official 'Total Capacity Under Study' directly from the report workbook."""
    _assert_ercot_url(report_url, "ERCOT report URL")
    content, content_type, effective_url = _fetch_url(report_url)

    workbook_bytes = _extract_workbook_bytes(content, effective_url, content_type)
    if workbook_bytes is None:
        return None

    return _extract_under_study_from_workbook(workbook_bytes)


def _fetch_url(url: str, timeout: int = REQUEST_TIMEOUT_SECONDS) -> tuple[bytes, str, str]:
    """Fetches URL content using curl first for ERCOT domains, or requests with fallback."""
    headers = {"User-Agent": "Mozilla/5.0"}
    is_ercot = _is_ercot_url(url)

    # Use curl immediately for ERCOT domains to avoid requests SSL handshake issues
    if is_ercot:
        try:
            output = _curl_fetch(url, headers["User-Agent"], timeout=timeout, try_legacy_cipher=False)
            if output:
                return output, "", url
        except Exception:
            try:
                output = _curl_fetch(url, headers["User-Agent"], timeout=timeout, try_legacy_cipher=True)
                if output:
                    return output, "", url
            except Exception:
                pass

    try:
        response = requests.get(
            url,
            headers=headers,
            timeout=timeout,
            allow_redirects=True,
            verify=False,
        )
        response.raise_for_status()
        content_type = response.headers.get("content-type", "").lower()
        return response.content, content_type, response.url
    except Exception as request_exc:
        # If not ERCOT and requests failed, try curl as fallback
        if not is_ercot:
            try:
                output = _curl_fetch(url, headers["User-Agent"], timeout=timeout, try_legacy_cipher=False)
                return output, "", url
            except Exception as curl_exc:
                raise RuntimeError(
                    f"Request failed via requests ({request_exc}) and curl fallback ({curl_exc})"
                ) from request_exc
        raise request_exc


def _extract_workbook_bytes(content: bytes, source_name: str, content_type: str) -> bytes | None:
    source_lower = source_name.lower()
    content_type_lower = (content_type or "").lower()

    if source_lower.endswith(".xlsx") or "spreadsheetml" in content_type_lower:
        return content

    # openpyxl cannot open legacy .xls files.
    if source_lower.endswith(".xls"):
        return None

    if source_lower.endswith(".zip") or "zip" in content_type_lower:
        return _extract_workbook_from_zip(content)

    # PK header can be either xlsx or zip. Try workbook first, then zip container.
    if content[:2] == b"PK":
        try:
            load_workbook(io.BytesIO(content), data_only=True, read_only=True)
            return content
        except Exception:
            return _extract_workbook_from_zip(content)

    return None


def _extract_workbook_from_zip(content: bytes) -> bytes | None:
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as zf:
            workbook_names = [
                name for name in zf.namelist() if name.lower().endswith(".xlsx")
            ]
            if not workbook_names:
                return None

            workbook_names.sort(
                key=lambda name: (
                    0 if "gis" in name.lower() or "report" in name.lower() else 1,
                    len(name),
                )
            )
            with zf.open(workbook_names[0]) as file_handle:
                return file_handle.read()
    except Exception:
        return None


def _extract_under_study_from_workbook(workbook_bytes: bytes) -> float | None:
    pattern = re.compile(r"total\s+capacity\s+under\s+study", flags=re.IGNORECASE)
    candidates: list[float] = []

    try:
        workbook = load_workbook(io.BytesIO(workbook_bytes), data_only=True, read_only=True)
    except Exception:
        return None

    sheet_names = list(workbook.sheetnames)
    preferred = [name for name in sheet_names if "summary" in name.lower()]
    ordered_sheet_names = preferred + [name for name in sheet_names if name not in preferred]

    for sheet_name in ordered_sheet_names:
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(values_only=True):
            if not row:
                continue
            if not any(pattern.search(str(value)) for value in row if isinstance(value, str)):
                continue

            for value in row:
                numeric = _to_positive_float(value)
                if numeric is not None:
                    candidates.append(numeric)

            if candidates and "summary" in sheet_name.lower():
                workbook.close()
                return max(candidates)

    workbook.close()
    return max(candidates) if candidates else None


def _to_positive_float(value: Any) -> float | None:
    if value is None:
        return None

    if isinstance(value, (int, float)):
        numeric = float(value)
        return numeric if numeric > 0 else None

    if isinstance(value, str):
        cleaned = re.sub(r"[^0-9.\-]", "", value)
        if not cleaned:
            return None
        try:
            numeric = float(cleaned)
            return numeric if numeric > 0 else None
        except ValueError:
            return None

    return None


def _curl_fetch(url: str, user_agent: str, *, timeout: int, try_legacy_cipher: bool) -> bytes:
    cmd = ["curl", "-k", "-L", "-sS"]
    if try_legacy_cipher:
        cmd.extend(["--ciphers", "DEFAULT@SECLEVEL=1"])
    cmd.extend(["-A", user_agent, url])
    return subprocess.check_output(cmd, timeout=timeout + 5)


def _extract_report_candidates(html: str, base_url: str) -> list[dict[str, Any]]:
    soup = BeautifulSoup(html, "html.parser")
    results: list[dict[str, Any]] = []
    seen: set[str] = set()

    for anchor in soup.find_all("a"):
        _process_anchor(anchor, base_url, seen, results)

    return results


def _process_anchor(
    anchor: Any, 
    base_url: str, 
    seen: set[str], 
    results: list[dict[str, Any]],
) -> None:
    href = (anchor.get("href") or "").strip()
    link_text = " ".join(anchor.get_text(" ", strip=True).split())
    
    if not href or href.startswith(("#", "javascript:", "mailto:")):
        return

    abs_url = urljoin(base_url, href)
    normalized = abs_url.lower()

    context_text = _derive_anchor_context(anchor)
    label = " ".join(part for part in (context_text, link_text) if part).strip()
    label_lower = label.lower()

    if "reporttypeid=" in normalized:
        return

    if normalized in seen:
        return

    if not _looks_like_report_link(normalized, label_lower):
        return

    score = _score_candidate(normalized, label_lower)
    found_date = _parse_date(normalized) or _parse_date(label)
    
    results.append(
        {
            "url": abs_url,
            "label": label,  # Store full context label for debugging
            "score": score,
            "date": found_date or datetime.min,
            "position": len(results), # use simple counter via list len
        }
    )
    seen.add(normalized)


def _derive_anchor_context(anchor: Any) -> str:
    contexts: list[str] = []

    for tr in anchor.find_parents("tr"):
        tr_text = " ".join(tr.get_text(" ", strip=True).split())
        if tr_text:
            contexts.append(tr_text)

    if contexts:
        # Use the smallest enclosing row to avoid grabbing entire table text.
        shortest = min(contexts, key=len)
        return shortest[:600]

    parent = anchor.parent
    if parent is not None:
        parent_text = " ".join(parent.get_text(" ", strip=True).split())
        if parent_text:
            return parent_text[:600]

    return ""


def _looks_like_report_link(url_lower: str, label_lower: str) -> bool:
    if any(url_lower.endswith(ext) for ext in DATA_EXTENSIONS):
        return True

    combined = f"{url_lower} {label_lower}"
    keywords = (
        "gis",
        "interconnection",
        "queue",
        "status report",
        "doclookupid",
        "getreport",
        "mirdownload",
    )
    if any(keyword in combined for keyword in keywords):
        return True

    if label_lower in ("xlsx", "xls", "csv", "zip") and "doclookupid=" in url_lower:
        return True

    return False


def _score_candidate(url_lower: str, label_lower: str) -> int:
    score = 0
    combined = f"{url_lower} {label_lower}"
    if "gis" in combined:
        score += 20  # Strong preference for GIS report
    if "interconnection" in combined or "queue" in combined:
        score += 3
    if "status" in combined:
        score += 1
    if "doclookupid" in url_lower or "getreport" in url_lower or "mirdownload" in url_lower:
        score += 4
    if any(url_lower.endswith(ext) for ext in DATA_EXTENSIONS):
        score += 2
    
    # Penalize URLs that look like index/discovery pages or unwanted reports
    if "reportid=" in url_lower or "reporttypeid=" in url_lower or "showhtmlview" in url_lower:
        score -= 20
    
    if "battery" in combined or "co-located" in combined:
        score -= 10

    return score


def _parse_date(value: str) -> datetime | None:
    month_match = MONTH_NAME_PATTERN.search(value)
    if month_match:
        month_name, year_text = month_match.groups()
        month = MONTH_TO_INT.get(month_name[:3].lower())
        if month:
            return datetime(int(year_text), month, 1)

    for pattern in DATE_PATTERNS:
        match = pattern.search(value)
        if not match:
            continue

        parts = [int(part) for part in match.groups()]
        if len(parts) == 3 and parts[0] > 1900 and parts[1] <= 12 and parts[2] <= 31:
            year, month, day = parts
        elif len(parts) == 3:
            month, day, year = parts
        else:
            continue

        try:
            return datetime(year, month, day)
        except ValueError:
            continue

    return None


def _download_table(url: str, depth: int = 0) -> pd.DataFrame:
    if depth > 2:
        raise RuntimeError("Exceeded redirect/discovery depth while fetching report data")

    content, content_type, effective_url = _fetch_url(url)
    url_lower = effective_url.lower()

    if _looks_like_html(content, content_type):
        html_text = content.decode("utf-8", errors="ignore")

        try:
            tables = pd.read_html(io.StringIO(html_text))
            if tables:
                largest_table = max(tables, key=len)
                if len(largest_table) > 0:
                    return largest_table
        except ValueError:
            pass

        nested_candidates = _extract_report_candidates(html_text, effective_url)
        nested_candidates = [candidate for candidate in nested_candidates if _is_ercot_url(str(candidate["url"]))]
        if nested_candidates:
            best_nested = _select_best_candidate(nested_candidates)
            return _download_table(str(best_nested["url"]), depth=depth + 1)

        raise RuntimeError(f"URL did not return tabular data: {effective_url}")

    # Try Excel first (since xlsx is also a zip file with PK header)
    try:
        df = _read_excel_with_fallbacks(content)
        if not df.empty:
            return df
    except Exception:
        pass

    if url_lower.endswith(".zip") or "zip" in content_type or content[:2] == b"PK":
        return _read_zip_table(content)

    return _read_table_bytes(content, effective_url, content_type)


def _read_zip_table(content: bytes) -> pd.DataFrame:
    with zipfile.ZipFile(io.BytesIO(content)) as zf:
        names = sorted(zf.namelist())
        # print(f"DEBUG ZIP CONTENT: {names}")
        for name in names:
            name_lower = name.lower()
            if not any(name_lower.endswith(ext) for ext in (".csv", ".xls", ".xlsx")):
                continue
            with zf.open(name) as file_handle:
                return _read_table_bytes(file_handle.read(), name, "")

    raise RuntimeError("ZIP file did not contain CSV/XLS/XLSX data")


def _read_table_bytes(content: bytes, source_name: str, content_type: str) -> pd.DataFrame:
    source_lower = source_name.lower()

    if source_lower.endswith((".xls", ".xlsx")) or "excel" in content_type or "spreadsheet" in content_type:
        return _read_excel_with_fallbacks(content)

    if source_lower.endswith(".csv") or "csv" in content_type or "text/plain" in content_type:
        return _read_csv_with_fallbacks(content)

    try:
        return _read_excel_with_fallbacks(content)
    except Exception:  # pylint: disable=broad-except
        return _read_csv_with_fallbacks(content)


def _read_excel_with_fallbacks(content: bytes) -> pd.DataFrame:
    xls = pd.ExcelFile(io.BytesIO(content))
    all_dataframes: list[pd.DataFrame] = []

    for sheet_name in xls.sheet_names:
        sheet_df, header_row, header_score = _read_sheet_best_effort(xls, sheet_name)
        if sheet_df.empty:
            continue

        cleaned = sheet_df.dropna(axis=1, how="all").dropna(axis=0, how="all")
        if cleaned.empty:
            continue

        cleaned["_source_sheet"] = sheet_name
        cleaned["_source_header_row"] = header_row
        cleaned["_source_header_score"] = header_score
        all_dataframes.append(cleaned)

    if not all_dataframes:
        return pd.DataFrame()

    # Consolidate all parsed sheets (all tabs processed)
    return pd.concat(all_dataframes, ignore_index=True, sort=False)


def _read_sheet_best_effort(xls: pd.ExcelFile, sheet_name: str) -> tuple[pd.DataFrame, int, int]:
    best_df = pd.DataFrame()
    best_score = -1
    best_header_row = 0

    for header_row in range(MAX_HEADER_SCAN_ROWS):
        try:
            df = pd.read_excel(xls, sheet_name=sheet_name, header=header_row)
        except Exception:
            continue

        if df is None or df.empty:
            continue

        cleaned = df.dropna(axis=1, how="all").dropna(axis=0, how="all")
        if cleaned.empty:
            continue

        header_score = _score_table_columns(cleaned.columns)
        width_bonus = min(len(cleaned.columns), 40)
        row_bonus = min(len(cleaned), 500) // 50
        score = header_score * 100 + width_bonus + row_bonus

        if score > best_score:
            best_score = score
            best_df = cleaned
            best_header_row = header_row

    if not best_df.empty:
        return best_df, best_header_row, best_score

    try:
        fallback = pd.read_excel(xls, sheet_name=sheet_name, header=0)
        fallback = fallback.dropna(axis=1, how="all").dropna(axis=0, how="all")
        if fallback.empty:
            return pd.DataFrame(), 0, -1
        return fallback, 0, _score_table_columns(fallback.columns)
    except Exception:
        return pd.DataFrame(), 0, -1


def _score_table_columns(columns: Any) -> int:
    score = 0
    column_text = " ".join(str(col).lower() for col in columns)

    if ("queue" in column_text and "id" in column_text) or "inr" in column_text or "gim" in column_text:
        score += 10
    if "project" in column_text or "name" in column_text:
        score += 6
    if "status" in column_text:
        score += 5
    if "mw" in column_text or "capacity" in column_text:
        score += 5
    if "county" in column_text:
        score += 4
    if "fuel" in column_text or "technology" in column_text:
        score += 4

    return score


def _is_gis_candidate(candidate: dict[str, Any]) -> bool:
    url_text = str(candidate.get("url", "")).lower()
    label_text = str(candidate.get("label", "")).lower()
    combined = f"{url_text} {label_text}"
    return (
        "gis_report" in combined
        or "gis report" in combined
        or "generation interconnection status" in combined
        or bool(re.search(r"\bgis\b", combined))
    )


def _select_best_candidate(candidates: list[dict[str, Any]]) -> dict[str, Any]:
    gis_candidates = [candidate for candidate in candidates if _is_gis_candidate(candidate)]
    pool = gis_candidates if gis_candidates else candidates
    ranked = sorted(pool, key=lambda item: (item["date"], item["score"], -item["position"]), reverse=True)
    return ranked[0]


def _read_csv_with_fallbacks(content: bytes) -> pd.DataFrame:
    errors: list[str] = []
    for encoding in ("utf-8", "utf-8-sig", "latin-1"):
        try:
            return pd.read_csv(io.BytesIO(content), encoding=encoding, low_memory=False)
        except Exception as exc:  # pylint: disable=broad-except
            errors.append(f"{encoding}: {exc}")

    raise RuntimeError("Could not parse CSV data. " + " | ".join(errors))


def _attach_tab_metadata(df: pd.DataFrame, base_meta: dict[str, Any]) -> dict[str, Any]:
    meta = dict(base_meta)
    source_col = "_source_sheet"
    if source_col in df.columns:
        tabs = sorted({str(tab) for tab in df[source_col].dropna().astype(str)})
        meta["tab_count"] = len(tabs)
        meta["tabs_processed"] = tabs
    return meta


def _assert_ercot_url(url: str, label: str) -> None:
    if not _is_ercot_url(url):
        raise ValueError(f"{label} must be hosted on an ERCOT domain (*.ercot.com): {url}")


def _is_ercot_url(url: str) -> bool:
    parsed = urlparse(url)
    host = (parsed.netloc or "").lower()
    return host.endswith("ercot.com")


def _looks_like_html(content: bytes, content_type: str) -> bool:
    if "html" in content_type:
        return True

    trimmed = content.lstrip()
    if not trimmed:
        return False
        
    lower_start = trimmed[:100].lower()
    return lower_start.startswith((b"<!doctype html", b"<html"))
